"""
Burnout Prediction — Scenario Runner + Excel Report
====================================================
Runs 20 realistic student profiles through all three strategy tiers
(Rule-Based, Hybrid, ML) and exports a colour-coded Excel workbook.

Run from project root:
    python backend/src/burnout_prediction/run_scenarios.py
"""

import sys
import os
import tempfile
import sqlite3
from pathlib import Path
from datetime import datetime, timedelta

# ── path setup ────────────────────────────────────────────────────────────────
_HERE = Path(__file__).resolve().parent
_SRC  = _HERE.parent
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))

from burnout_prediction.schemas import FinancialFeatures, MentalFeatures, BurnoutAlertLevel
from burnout_prediction.strategies.rule_based import RuleBasedStrategy
from burnout_prediction.strategies.hybrid   import HybridStrategy
from burnout_prediction.strategies.ml_strategy import MLStrategy
from burnout_prediction.score_combiner import ScoreCombiner
from burnout_prediction.feature_engineer import FeatureEngineer

# ── 20 scenarios ──────────────────────────────────────────────────────────────
# Each entry: (label, FinancialFeatures, MentalFeatures)
# Profiles cover the full spectrum from thriving to crisis.

SCENARIOS = [

    # ── GOOD zone (combined < 0.30) ──────────────────────────────────────────
    (
        "1. Thriving Student",
        FinancialFeatures(
            daily_spend_avg_7d=12, daily_spend_avg_30d=11,
            weekly_spend_current=84, weekly_spend_last=80,
            expense_growth_rate=0.05, projected_month_end_spend=360,
            budget_utilization_pct=40, days_until_broke=30,
            food_spend_ratio=0.30, transport_spend_ratio=0.10,
            entertainment_spend_ratio=0.05,
            impulse_spend_count=0, budget_safety_margin=540,
            category_overspend_count=0, food_homemade_ratio=0.85,
            skipped_meals_count=0, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=8.0, sleep_deficit=0.0,
            stress_avg_7d=2.5, stress_trend=-0.1,
            exercise_avg_7d=45, no_exercise_days_7d=1,
            mood_score_avg=0.85, bad_mood_streak=0,
            energy_avg_7d=8.0, social_activity_avg_7d=6.0,
            goal_completion_rate=0.95, missed_goals_count=0,
            skipped_meals_count=0, sleep_consistency=0.3,
            stress_sleep_interaction=0.5,
        ),
    ),
    (
        "2. Budget-Conscious, Calm",
        FinancialFeatures(
            daily_spend_avg_7d=15, daily_spend_avg_30d=14,
            weekly_spend_current=105, weekly_spend_last=98,
            expense_growth_rate=0.07, projected_month_end_spend=450,
            budget_utilization_pct=50, days_until_broke=28,
            food_spend_ratio=0.35, transport_spend_ratio=0.15,
            entertainment_spend_ratio=0.10,
            impulse_spend_count=1, budget_safety_margin=450,
            category_overspend_count=0, food_homemade_ratio=0.70,
            skipped_meals_count=1, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=7.5, sleep_deficit=0.0,
            stress_avg_7d=3.5, stress_trend=0.0,
            exercise_avg_7d=30, no_exercise_days_7d=2,
            mood_score_avg=0.60, bad_mood_streak=0,
            energy_avg_7d=7.0, social_activity_avg_7d=4.0,
            goal_completion_rate=0.80, missed_goals_count=0,
            skipped_meals_count=1, sleep_consistency=0.5,
            stress_sleep_interaction=0.9,
        ),
    ),
    (
        "3. Stable with Minor Stress",
        FinancialFeatures(
            daily_spend_avg_7d=18, daily_spend_avg_30d=16,
            weekly_spend_current=126, weekly_spend_last=115,
            expense_growth_rate=0.10, projected_month_end_spend=520,
            budget_utilization_pct=58, days_until_broke=22,
            food_spend_ratio=0.38, transport_spend_ratio=0.18,
            entertainment_spend_ratio=0.12,
            impulse_spend_count=2, budget_safety_margin=380,
            category_overspend_count=1, food_homemade_ratio=0.60,
            skipped_meals_count=2, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=7.0, sleep_deficit=0.3,
            stress_avg_7d=4.5, stress_trend=0.1,
            exercise_avg_7d=20, no_exercise_days_7d=3,
            mood_score_avg=0.40, bad_mood_streak=0,
            energy_avg_7d=6.0, social_activity_avg_7d=3.5,
            goal_completion_rate=0.75, missed_goals_count=1,
            skipped_meals_count=2, sleep_consistency=0.6,
            stress_sleep_interaction=1.4,
        ),
    ),

    # ── MODERATE zone (0.30 – 0.55) ──────────────────────────────────────────
    (
        "4. Mild Overspending",
        FinancialFeatures(
            daily_spend_avg_7d=28, daily_spend_avg_30d=22,
            weekly_spend_current=196, weekly_spend_last=160,
            expense_growth_rate=0.23, projected_month_end_spend=750,
            budget_utilization_pct=78, days_until_broke=15,
            food_spend_ratio=0.42, transport_spend_ratio=0.20,
            entertainment_spend_ratio=0.18,
            impulse_spend_count=3, budget_safety_margin=150,
            category_overspend_count=2, food_homemade_ratio=0.45,
            skipped_meals_count=2, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=6.8, sleep_deficit=0.5,
            stress_avg_7d=5.0, stress_trend=0.15,
            exercise_avg_7d=15, no_exercise_days_7d=4,
            mood_score_avg=0.25, bad_mood_streak=1,
            energy_avg_7d=5.5, social_activity_avg_7d=3.0,
            goal_completion_rate=0.65, missed_goals_count=1,
            skipped_meals_count=2, sleep_consistency=0.8,
            stress_sleep_interaction=1.7,
        ),
    ),
    (
        "5. Anxious + Budget Creep",
        FinancialFeatures(
            daily_spend_avg_7d=32, daily_spend_avg_30d=25,
            weekly_spend_current=224, weekly_spend_last=185,
            expense_growth_rate=0.21, projected_month_end_spend=850,
            budget_utilization_pct=85, days_until_broke=12,
            food_spend_ratio=0.44, transport_spend_ratio=0.22,
            entertainment_spend_ratio=0.14,
            impulse_spend_count=3, budget_safety_margin=50,
            category_overspend_count=2, food_homemade_ratio=0.40,
            skipped_meals_count=3, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=6.2, sleep_deficit=0.8,
            stress_avg_7d=6.0, stress_trend=0.3,
            exercise_avg_7d=10, no_exercise_days_7d=5,
            mood_score_avg=-0.10, bad_mood_streak=2,
            energy_avg_7d=4.5, social_activity_avg_7d=2.5,
            goal_completion_rate=0.50, missed_goals_count=2,
            skipped_meals_count=3, sleep_consistency=1.1,
            stress_sleep_interaction=2.4,
        ),
    ),
    (
        "6. Exam Week — Mentally Stressed",
        FinancialFeatures(
            daily_spend_avg_7d=20, daily_spend_avg_30d=18,
            weekly_spend_current=140, weekly_spend_last=135,
            expense_growth_rate=0.04, projected_month_end_spend=580,
            budget_utilization_pct=60, days_until_broke=20,
            food_spend_ratio=0.36, transport_spend_ratio=0.14,
            entertainment_spend_ratio=0.08,
            impulse_spend_count=1, budget_safety_margin=320,
            category_overspend_count=1, food_homemade_ratio=0.55,
            skipped_meals_count=3, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=5.5, sleep_deficit=1.5,
            stress_avg_7d=7.0, stress_trend=0.5,
            exercise_avg_7d=5, no_exercise_days_7d=5,
            mood_score_avg=-0.20, bad_mood_streak=2,
            energy_avg_7d=4.0, social_activity_avg_7d=1.5,
            goal_completion_rate=0.45, missed_goals_count=3,
            skipped_meals_count=4, sleep_consistency=1.3,
            stress_sleep_interaction=3.5,
        ),
    ),
    (
        "7. Financially Stressed, OK Mentally",
        FinancialFeatures(
            daily_spend_avg_7d=42, daily_spend_avg_30d=35,
            weekly_spend_current=294, weekly_spend_last=245,
            expense_growth_rate=0.20, projected_month_end_spend=1100,
            budget_utilization_pct=95, days_until_broke=8,
            food_spend_ratio=0.48, transport_spend_ratio=0.25,
            entertainment_spend_ratio=0.15,
            impulse_spend_count=5, budget_safety_margin=-100,
            category_overspend_count=3, food_homemade_ratio=0.25,
            skipped_meals_count=2, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=7.2, sleep_deficit=0.1,
            stress_avg_7d=4.8, stress_trend=-0.1,
            exercise_avg_7d=25, no_exercise_days_7d=2,
            mood_score_avg=0.30, bad_mood_streak=0,
            energy_avg_7d=6.5, social_activity_avg_7d=4.0,
            goal_completion_rate=0.80, missed_goals_count=0,
            skipped_meals_count=1, sleep_consistency=0.5,
            stress_sleep_interaction=1.4,
        ),
    ),
    (
        "8. Skipping Meals + Low Energy",
        FinancialFeatures(
            daily_spend_avg_7d=22, daily_spend_avg_30d=20,
            weekly_spend_current=154, weekly_spend_last=140,
            expense_growth_rate=0.10, projected_month_end_spend=620,
            budget_utilization_pct=65, days_until_broke=18,
            food_spend_ratio=0.30, transport_spend_ratio=0.12,
            entertainment_spend_ratio=0.10,
            impulse_spend_count=2, budget_safety_margin=280,
            category_overspend_count=1, food_homemade_ratio=0.20,
            skipped_meals_count=5, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=6.0, sleep_deficit=1.0,
            stress_avg_7d=6.2, stress_trend=0.2,
            exercise_avg_7d=8, no_exercise_days_7d=5,
            mood_score_avg=-0.15, bad_mood_streak=2,
            energy_avg_7d=3.5, social_activity_avg_7d=2.0,
            goal_completion_rate=0.40, missed_goals_count=3,
            skipped_meals_count=5, sleep_consistency=1.4,
            stress_sleep_interaction=2.5,
        ),
    ),

    # ── HIGH zone (0.55 – 0.75) ──────────────────────────────────────────────
    (
        "9. Near Broke, Overwhelmed",
        FinancialFeatures(
            daily_spend_avg_7d=50, daily_spend_avg_30d=38,
            weekly_spend_current=350, weekly_spend_last=280,
            expense_growth_rate=0.25, projected_month_end_spend=1200,
            budget_utilization_pct=105, days_until_broke=6,
            food_spend_ratio=0.50, transport_spend_ratio=0.25,
            entertainment_spend_ratio=0.15,
            impulse_spend_count=6, budget_safety_margin=-300,
            category_overspend_count=4, food_homemade_ratio=0.10,
            skipped_meals_count=4, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=5.8, sleep_deficit=1.5,
            stress_avg_7d=7.5, stress_trend=0.5,
            exercise_avg_7d=5, no_exercise_days_7d=6,
            mood_score_avg=-0.45, bad_mood_streak=3,
            energy_avg_7d=3.0, social_activity_avg_7d=1.5,
            goal_completion_rate=0.25, missed_goals_count=4,
            skipped_meals_count=4, sleep_consistency=1.8,
            stress_sleep_interaction=3.6,
        ),
    ),
    (
        "10. Spending Surge + Sleep Debt",
        FinancialFeatures(
            daily_spend_avg_7d=55, daily_spend_avg_30d=35,
            weekly_spend_current=385, weekly_spend_last=245,
            expense_growth_rate=0.57, projected_month_end_spend=1400,
            budget_utilization_pct=108, days_until_broke=5,
            food_spend_ratio=0.52, transport_spend_ratio=0.22,
            entertainment_spend_ratio=0.18,
            impulse_spend_count=7, budget_safety_margin=-500,
            category_overspend_count=4, food_homemade_ratio=0.08,
            skipped_meals_count=4, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=5.2, sleep_deficit=2.0,
            stress_avg_7d=7.8, stress_trend=0.6,
            exercise_avg_7d=3, no_exercise_days_7d=6,
            mood_score_avg=-0.50, bad_mood_streak=3,
            energy_avg_7d=2.8, social_activity_avg_7d=1.0,
            goal_completion_rate=0.20, missed_goals_count=4,
            skipped_meals_count=5, sleep_consistency=2.0,
            stress_sleep_interaction=4.0,
        ),
    ),
    (
        "11. Month-End Panic",
        FinancialFeatures(
            daily_spend_avg_7d=60, daily_spend_avg_30d=42,
            weekly_spend_current=420, weekly_spend_last=300,
            expense_growth_rate=0.40, projected_month_end_spend=1350,
            budget_utilization_pct=110, days_until_broke=4,
            food_spend_ratio=0.55, transport_spend_ratio=0.20,
            entertainment_spend_ratio=0.15,
            impulse_spend_count=7, budget_safety_margin=-450,
            category_overspend_count=5, food_homemade_ratio=0.10,
            skipped_meals_count=4, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=6.0, sleep_deficit=1.2,
            stress_avg_7d=7.0, stress_trend=0.4,
            exercise_avg_7d=5, no_exercise_days_7d=5,
            mood_score_avg=-0.40, bad_mood_streak=3,
            energy_avg_7d=3.5, social_activity_avg_7d=1.5,
            goal_completion_rate=0.20, missed_goals_count=4,
            skipped_meals_count=4, sleep_consistency=1.7,
            stress_sleep_interaction=3.0,
        ),
    ),
    (
        "12. Depressed but Financially OK",
        FinancialFeatures(
            daily_spend_avg_7d=16, daily_spend_avg_30d=15,
            weekly_spend_current=112, weekly_spend_last=108,
            expense_growth_rate=0.04, projected_month_end_spend=470,
            budget_utilization_pct=52, days_until_broke=26,
            food_spend_ratio=0.32, transport_spend_ratio=0.12,
            entertainment_spend_ratio=0.06,
            impulse_spend_count=0, budget_safety_margin=430,
            category_overspend_count=0, food_homemade_ratio=0.75,
            skipped_meals_count=5, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=4.8, sleep_deficit=2.2,
            stress_avg_7d=8.0, stress_trend=0.7,
            exercise_avg_7d=0, no_exercise_days_7d=7,
            mood_score_avg=-0.65, bad_mood_streak=5,
            energy_avg_7d=2.5, social_activity_avg_7d=0.5,
            goal_completion_rate=0.10, missed_goals_count=5,
            skipped_meals_count=5, sleep_consistency=2.2,
            stress_sleep_interaction=4.4,
        ),
    ),
    (
        "13. High Transport Costs, Exhausted",
        FinancialFeatures(
            daily_spend_avg_7d=45, daily_spend_avg_30d=38,
            weekly_spend_current=315, weekly_spend_last=270,
            expense_growth_rate=0.17, projected_month_end_spend=1100,
            budget_utilization_pct=98, days_until_broke=7,
            food_spend_ratio=0.28, transport_spend_ratio=0.45,
            entertainment_spend_ratio=0.12,
            impulse_spend_count=4, budget_safety_margin=-100,
            category_overspend_count=3, food_homemade_ratio=0.60,
            skipped_meals_count=2, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=5.8, sleep_deficit=1.5,
            stress_avg_7d=7.2, stress_trend=0.3,
            exercise_avg_7d=10, no_exercise_days_7d=5,
            mood_score_avg=-0.30, bad_mood_streak=3,
            energy_avg_7d=3.8, social_activity_avg_7d=2.0,
            goal_completion_rate=0.30, missed_goals_count=3,
            skipped_meals_count=3, sleep_consistency=1.6,
            stress_sleep_interaction=3.2,
        ),
    ),

    # ── CRISIS zone (> 0.75) ─────────────────────────────────────────────────
    (
        "14. Full Financial + Mental Collapse",
        FinancialFeatures(
            daily_spend_avg_7d=55, daily_spend_avg_30d=40,
            weekly_spend_current=385, weekly_spend_last=220,
            expense_growth_rate=0.75, projected_month_end_spend=1700,
            budget_utilization_pct=120, days_until_broke=4,
            food_spend_ratio=0.55, transport_spend_ratio=0.20,
            entertainment_spend_ratio=0.15,
            impulse_spend_count=8, budget_safety_margin=-700,
            category_overspend_count=5, food_homemade_ratio=0.05,
            skipped_meals_count=5, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=4.2, sleep_deficit=2.8,
            stress_avg_7d=8.5, stress_trend=0.8,
            exercise_avg_7d=2, no_exercise_days_7d=6,
            mood_score_avg=-0.70, bad_mood_streak=4,
            energy_avg_7d=2.5, social_activity_avg_7d=1.0,
            goal_completion_rate=0.15, missed_goals_count=4,
            skipped_meals_count=5, sleep_consistency=2.5,
            stress_sleep_interaction=7.4,
        ),
    ),
    (
        "15. Zero Budget Left, Overwhelmed",
        FinancialFeatures(
            daily_spend_avg_7d=65, daily_spend_avg_30d=50,
            weekly_spend_current=455, weekly_spend_last=350,
            expense_growth_rate=0.30, projected_month_end_spend=1800,
            budget_utilization_pct=130, days_until_broke=2,
            food_spend_ratio=0.50, transport_spend_ratio=0.25,
            entertainment_spend_ratio=0.15,
            impulse_spend_count=9, budget_safety_margin=-900,
            category_overspend_count=5, food_homemade_ratio=0.05,
            skipped_meals_count=6, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=4.0, sleep_deficit=3.0,
            stress_avg_7d=9.0, stress_trend=0.9,
            exercise_avg_7d=0, no_exercise_days_7d=7,
            mood_score_avg=-0.90, bad_mood_streak=6,
            energy_avg_7d=2.0, social_activity_avg_7d=0.5,
            goal_completion_rate=0.05, missed_goals_count=5,
            skipped_meals_count=6, sleep_consistency=2.8,
            stress_sleep_interaction=8.1,
        ),
    ),
    (
        "16. Chronic Burnout",
        FinancialFeatures(
            daily_spend_avg_7d=48, daily_spend_avg_30d=45,
            weekly_spend_current=336, weekly_spend_last=315,
            expense_growth_rate=0.07, projected_month_end_spend=1350,
            budget_utilization_pct=115, days_until_broke=3,
            food_spend_ratio=0.48, transport_spend_ratio=0.22,
            entertainment_spend_ratio=0.18,
            impulse_spend_count=8, budget_safety_margin=-450,
            category_overspend_count=5, food_homemade_ratio=0.08,
            skipped_meals_count=6, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=4.5, sleep_deficit=2.5,
            stress_avg_7d=8.8, stress_trend=0.5,
            exercise_avg_7d=1, no_exercise_days_7d=7,
            mood_score_avg=-0.80, bad_mood_streak=7,
            energy_avg_7d=2.0, social_activity_avg_7d=0.5,
            goal_completion_rate=0.10, missed_goals_count=5,
            skipped_meals_count=6, sleep_consistency=2.6,
            stress_sleep_interaction=7.0,
        ),
    ),

    # ── Edge / Interesting cases ──────────────────────────────────────────────
    (
        "17. Rich but Mentally Broken",
        FinancialFeatures(
            daily_spend_avg_7d=10, daily_spend_avg_30d=9,
            weekly_spend_current=70, weekly_spend_last=68,
            expense_growth_rate=0.03, projected_month_end_spend=300,
            budget_utilization_pct=30, days_until_broke=30,
            food_spend_ratio=0.28, transport_spend_ratio=0.10,
            entertainment_spend_ratio=0.05,
            impulse_spend_count=0, budget_safety_margin=600,
            category_overspend_count=0, food_homemade_ratio=0.90,
            skipped_meals_count=0, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=4.0, sleep_deficit=3.5,
            stress_avg_7d=9.5, stress_trend=1.0,
            exercise_avg_7d=0, no_exercise_days_7d=7,
            mood_score_avg=-0.95, bad_mood_streak=7,
            energy_avg_7d=1.5, social_activity_avg_7d=0.0,
            goal_completion_rate=0.0, missed_goals_count=5,
            skipped_meals_count=6, sleep_consistency=3.0,
            stress_sleep_interaction=9.5,
        ),
    ),
    (
        "18. Broke but Mentally Strong",
        FinancialFeatures(
            daily_spend_avg_7d=55, daily_spend_avg_30d=40,
            weekly_spend_current=385, weekly_spend_last=250,
            expense_growth_rate=0.54, projected_month_end_spend=1500,
            budget_utilization_pct=118, days_until_broke=3,
            food_spend_ratio=0.55, transport_spend_ratio=0.22,
            entertainment_spend_ratio=0.15,
            impulse_spend_count=9, budget_safety_margin=-600,
            category_overspend_count=5, food_homemade_ratio=0.05,
            skipped_meals_count=5, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=7.8, sleep_deficit=0.0,
            stress_avg_7d=3.0, stress_trend=-0.2,
            exercise_avg_7d=40, no_exercise_days_7d=1,
            mood_score_avg=0.70, bad_mood_streak=0,
            energy_avg_7d=7.5, social_activity_avg_7d=5.0,
            goal_completion_rate=0.90, missed_goals_count=0,
            skipped_meals_count=0, sleep_consistency=0.4,
            stress_sleep_interaction=0.8,
        ),
    ),
    (
        "19. Recovering — Was Bad, Getting Better",
        FinancialFeatures(
            daily_spend_avg_7d=25, daily_spend_avg_30d=38,
            weekly_spend_current=175, weekly_spend_last=265,
            expense_growth_rate=-0.34, projected_month_end_spend=680,
            budget_utilization_pct=72, days_until_broke=14,
            food_spend_ratio=0.40, transport_spend_ratio=0.18,
            entertainment_spend_ratio=0.10,
            impulse_spend_count=2, budget_safety_margin=220,
            category_overspend_count=1, food_homemade_ratio=0.65,
            skipped_meals_count=1, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=6.8, sleep_deficit=0.5,
            stress_avg_7d=5.5, stress_trend=-0.6,
            exercise_avg_7d=20, no_exercise_days_7d=3,
            mood_score_avg=0.15, bad_mood_streak=1,
            energy_avg_7d=5.5, social_activity_avg_7d=3.0,
            goal_completion_rate=0.60, missed_goals_count=1,
            skipped_meals_count=2, sleep_consistency=0.9,
            stress_sleep_interaction=1.9,
        ),
    ),
    (
        "20. End of Semester Crunch",
        FinancialFeatures(
            daily_spend_avg_7d=35, daily_spend_avg_30d=28,
            weekly_spend_current=245, weekly_spend_last=196,
            expense_growth_rate=0.25, projected_month_end_spend=900,
            budget_utilization_pct=88, days_until_broke=9,
            food_spend_ratio=0.45, transport_spend_ratio=0.20,
            entertainment_spend_ratio=0.10,
            impulse_spend_count=4, budget_safety_margin=0,
            category_overspend_count=3, food_homemade_ratio=0.30,
            skipped_meals_count=4, emergency_fund=100, spendable_budget=900,
        ),
        MentalFeatures(
            sleep_avg_7d=5.5, sleep_deficit=1.8,
            stress_avg_7d=7.8, stress_trend=0.7,
            exercise_avg_7d=5, no_exercise_days_7d=6,
            mood_score_avg=-0.50, bad_mood_streak=4,
            energy_avg_7d=3.0, social_activity_avg_7d=1.5,
            goal_completion_rate=0.20, missed_goals_count=4,
            skipped_meals_count=4, sleep_consistency=2.0,
            stress_sleep_interaction=5.4,
        ),
    ),
]

# ── Run through rule-based strategy (no DB needed) ────────────────────────────

def run_all_scenarios():
    rule   = RuleBasedStrategy()
    hybrid = HybridStrategy()
    combiner = ScoreCombiner()
    results = []

    for label, fin, men in SCENARIOS:
        # Days-of-data labels per tier
        for days, tag in [(2, "Rule-Based"), (5, "Hybrid")]:
            if tag == "Rule-Based":
                r = rule.predict(fin, men, days_of_data=days)
            else:
                r = hybrid.predict(fin, men, days_of_data=days)

            alert = combiner.get_alert_level(r.combined_score)
            recs  = combiner.build_recommendations(r, alert)

            results.append({
                "Scenario":          label,
                "Strategy":          tag,

                # ── Inputs: Financial ──
                "Daily Avg Spend 7d":        fin.daily_spend_avg_7d,
                "Daily Avg Spend 30d":       fin.daily_spend_avg_30d,
                "Week Current Spend":        fin.weekly_spend_current,
                "Week Last Spend":           fin.weekly_spend_last,
                "Expense Growth Rate":       f"{fin.expense_growth_rate*100:.0f}%",
                "Projected Month-End":       fin.projected_month_end_spend,
                "Budget Utilisation %":      fin.budget_utilization_pct,
                "Days Until Broke":          fin.days_until_broke,
                "Food Spend Ratio":          f"{fin.food_spend_ratio*100:.0f}%",
                "Transport Ratio":           f"{fin.transport_spend_ratio*100:.0f}%",
                "Entertainment Ratio":       f"{fin.entertainment_spend_ratio*100:.0f}%",
                "Impulse Spends":            fin.impulse_spend_count,
                "Budget Safety Margin":      fin.budget_safety_margin,
                "Cat Overspend Count":       fin.category_overspend_count,
                "Homemade Food Ratio":       f"{fin.food_homemade_ratio*100:.0f}%",
                "Skipped Meals (fin)":       fin.skipped_meals_count,

                # ── Inputs: Mental ──
                "Sleep Avg 7d (hrs)":        men.sleep_avg_7d,
                "Sleep Deficit (hrs)":       men.sleep_deficit,
                "Stress Avg 7d":             men.stress_avg_7d,
                "Stress Trend":              round(men.stress_trend, 2),
                "Exercise Avg (min)":        men.exercise_avg_7d,
                "No-Exercise Days":          men.no_exercise_days_7d,
                "Mood Score Avg":            round(men.mood_score_avg, 2),
                "Bad Mood Streak":           men.bad_mood_streak,
                "Energy Avg":                men.energy_avg_7d,
                "Social Activity Avg":       men.social_activity_avg_7d,
                "Goal Completion %":         f"{men.goal_completion_rate*100:.0f}%",
                "Missed Goals":              men.missed_goals_count,
                "Skipped Meals (men)":       men.skipped_meals_count,
                "Sleep Consistency (std)":   round(men.sleep_consistency, 2),
                "Stress×Sleep Interaction":  round(men.stress_sleep_interaction, 2),

                # ── Outputs ──
                "Financial Score":  round(r.financial_score, 3),
                "Mental Score":     round(r.mental_score, 3),
                "Combined Score":   round(r.combined_score, 3),
                "Alert Level":      alert.value.upper(),
                "Confidence":       round(r.confidence, 2),
                "Top Risk Factor":  r.risk_factors[0] if r.risk_factors else "—",
                "Recommendation 1": recs[0] if len(recs) > 0 else "—",
                "Recommendation 2": recs[1] if len(recs) > 1 else "—",
            })

    return results


# ── Build Excel workbook ───────────────────────────────────────────────────────

def build_excel(results, out_path):
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    ALERT_FILL = {
        "GOOD":     PatternFill("solid", fgColor="C6EFCE"),   # green
        "MODERATE": PatternFill("solid", fgColor="FFEB9C"),   # yellow
        "HIGH":     PatternFill("solid", fgColor="FFCC99"),   # orange
        "CRISIS":   PatternFill("solid", fgColor="FFC7CE"),   # red
    }
    SCORE_FILL = {
        "GOOD":     PatternFill("solid", fgColor="E2EFDA"),
        "MODERATE": PatternFill("solid", fgColor="FFF2CC"),
        "HIGH":     PatternFill("solid", fgColor="FCE4D6"),
        "CRISIS":   PatternFill("solid", fgColor="F4CCCC"),
    }
    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
    SECTION_FILL_FIN = PatternFill("solid", fgColor="DDEEFF")
    SECTION_FILL_MEN = PatternFill("solid", fgColor="E8F5E9")
    SECTION_FILL_OUT = PatternFill("solid", fgColor="FFF9C4")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Full Data ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Scenario Results"

    if not results:
        wb.save(out_path)
        return

    headers = list(results[0].keys())
    # Section header rows above the column names
    section_map = {
        "Scenario":                    ("INFO",        HEADER_FILL),
        "Strategy":                    ("INFO",        HEADER_FILL),
        "Daily Avg Spend 7d":          ("FINANCIAL INPUTS", SECTION_FILL_FIN),
        "Sleep Avg 7d (hrs)":          ("MENTAL INPUTS",    SECTION_FILL_MEN),
        "Financial Score":             ("OUTPUTS",          SECTION_FILL_OUT),
    }

    # Row 1: section labels
    ws.row_dimensions[1].height = 16
    current_section = ""
    section_start = 1
    for ci, h in enumerate(headers, 1):
        if h in section_map:
            current_section = section_map[h][0]
            section_start = ci
        ws.cell(1, ci).fill = section_map.get(h, (None, HEADER_FILL))[1]

    # Write section banners in merged cells manually
    sec_ranges = [
        ("INFO",             1,  2),
        ("FINANCIAL INPUTS", 3,  18),
        ("MENTAL INPUTS",   19,  33),
        ("OUTPUTS",         34,  len(headers)),
    ]
    for sec_label, c1, c2 in sec_ranges:
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        cell = ws.cell(1, c1)
        cell.value = sec_label
        cell.font  = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        fill_map = {
            "INFO":             PatternFill("solid", fgColor="1F3864"),
            "FINANCIAL INPUTS": PatternFill("solid", fgColor="2E5FA3"),
            "MENTAL INPUTS":    PatternFill("solid", fgColor="2E8B57"),
            "OUTPUTS":          PatternFill("solid", fgColor="B8860B"),
        }
        cell.fill = fill_map[sec_label]

    # Row 2: column headers
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(2, ci, h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[2].height = 36

    # Data rows start at row 3
    for ri, row in enumerate(results, 3):
        alert = row["Alert Level"]
        for ci, (key, val) in enumerate(row.items(), 1):
            cell = ws.cell(ri, ci, val)
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # colour score columns
            if key in ("Financial Score", "Mental Score", "Combined Score"):
                cell.fill = SCORE_FILL.get(alert, PatternFill())
                cell.font = Font(bold=True)
            elif key == "Alert Level":
                cell.fill = ALERT_FILL.get(alert, PatternFill())
                cell.font = Font(bold=True)
            elif key == "Scenario":
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif key in ("Recommendation 1", "Recommendation 2", "Top Risk Factor"):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.row_dimensions[ri].height = 28

    # Column widths
    col_widths = {
        1: 32, 2: 12,
        3: 14, 4: 14, 5: 13, 6: 13, 7: 14, 8: 16, 9: 14, 10: 13,
        11: 12, 12: 14, 13: 14, 14: 12, 15: 16, 16: 14, 17: 13, 18: 13,
        19: 14, 20: 13, 21: 11, 22: 12, 23: 14, 24: 13, 25: 12,
        26: 12, 27: 14, 28: 12, 29: 12, 30: 14, 31: 14, 32: 13, 33: 18,
        34: 14, 35: 14, 36: 14, 37: 12, 38: 11, 39: 36, 40: 44, 41: 44,
    }
    for ci, width in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.freeze_panes = "C3"

    # ── Sheet 2: Summary ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 36

    title_fill = PatternFill("solid", fgColor="1F3864")
    title_font = Font(color="FFFFFF", bold=True, size=12)

    headers2 = ["Scenario", "Strategy", "Fin Score", "Men Score",
                "Combined", "Alert Level", "Top Risk Factor"]
    for ci, h in enumerate(headers2, 1):
        c = ws2.cell(1, ci, h)
        c.fill = title_fill; c.font = title_font
        c.alignment = Alignment(horizontal="center")
        c.border = border
    ws2.row_dimensions[1].height = 24

    for ri, row in enumerate(results, 2):
        alert = row["Alert Level"]
        vals  = [row["Scenario"], row["Strategy"],
                 row["Financial Score"], row["Mental Score"],
                 row["Combined Score"], alert,
                 row["Top Risk Factor"]]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(ri, ci, v)
            c.border = border
            c.alignment = Alignment(horizontal="center" if ci != 7 else "left",
                                    vertical="center", wrap_text=True)
            if ci == 6:
                c.fill = ALERT_FILL.get(alert, PatternFill())
                c.font = Font(bold=True)
            if ci in (3, 4, 5):
                c.fill = SCORE_FILL.get(alert, PatternFill())
        ws2.row_dimensions[ri].height = 22

    # Alert level count table
    ws2.cell(len(results) + 4, 1, "Alert Level Distribution").font = Font(bold=True, size=11)
    for i, level in enumerate(["GOOD", "MODERATE", "HIGH", "CRISIS"], len(results) + 5):
        count = sum(1 for r in results if r["Alert Level"] == level)
        c = ws2.cell(i, 1, level)
        c.fill = ALERT_FILL[level]
        c.font = Font(bold=True)
        c.border = border
        ws2.cell(i, 2, count).border = border
        ws2.cell(i, 2).alignment = Alignment(horizontal="center")

    wb.save(out_path)
    print(f"\n  Excel saved → {out_path}")


# ── Console summary ───────────────────────────────────────────────────────────

def print_summary(results):
    rule_results    = [r for r in results if r["Strategy"] == "Rule-Based"]
    hybrid_results  = [r for r in results if r["Strategy"] == "Hybrid"]

    print("\n" + "="*80)
    print("  BURNOUT PREDICTION — SCENARIO RESULTS (Rule-Based Strategy)")
    print("="*80)
    print(f"  {'Scenario':<42} {'Fin':>6} {'Men':>6} {'Combined':>9}  {'Alert':<10}")
    print("-"*80)

    for r in rule_results:
        fin_s = r["Financial Score"]
        men_s = r["Mental Score"]
        com_s = r["Combined Score"]
        alert = r["Alert Level"]
        marker = {"GOOD": " ✓", "MODERATE": " ~", "HIGH": " !", "CRISIS": " ✗"}.get(alert, "")
        print(f"  {r['Scenario']:<42} {fin_s:>6.3f} {men_s:>6.3f} {com_s:>9.3f}  {alert:<10}{marker}")

    print("\n" + "="*80)
    print("  Distribution (Rule-Based)")
    print("-"*80)
    for level in ["GOOD", "MODERATE", "HIGH", "CRISIS"]:
        count = sum(1 for r in rule_results if r["Alert Level"] == level)
        bar   = "█" * count
        print(f"  {level:<10} {bar} ({count})")

    print("="*80)


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    results = run_all_scenarios()
    print_summary(results)

    out = Path(__file__).parent / "burnout_scenario_results.xlsx"
    build_excel(results, str(out))
    print(f"\n  Scenarios run : {len(SCENARIOS)}")
    print(f"  Total rows    : {len(results)} (each scenario × 2 strategies)")
    print(f"  Excel path    : {out}\n")
